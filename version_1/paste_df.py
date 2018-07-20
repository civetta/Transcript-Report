import pandas as pd
import re
from rt_df import rt_df
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def paste_df(df):
    """Creates a dataframe for each teacher, and is then called teacher_df
    It then creates a workbook, called teacherbook. Within teacherbook
    there are two worksheets. One titled 'Transcripts' the other titled
    'Response Times'. Then it calls the functions that paste the transcripts,
    the response times, and the statistics table"""
    teacher_real_names = df.name
    unique_teacher_names = list(set(teacher_real_names))
    for teachername in unique_teacher_names:
        teacher_df = df[(df.name == teachername)]
        teacher_df = teacher_df.reset_index(drop=True)
        teacherbook = Workbook()
        ws = teacherbook.create_sheet("Transcripts", 0) 
        for_each_transcript(teacher_df, teachername, ws, teacherbook)
        response_time_ws = teacherbook.create_sheet('Response Times',1)
        paste_time_stamps(teacher_df, response_time_ws)
        create_stats_table(df, teacher_df, response_time_ws)
        teacherbook.save('teacher_sheets/'+teachername+'.xlsx')


def for_each_transcript(teacher_df, teachername, ws, teacherbook):
    """With the new Teacher Dataframe, which only contains information relavtie
    to teachername, it goes through each row and calls each transcript. It then
    uses the whiteboard, and lesson title to make the first row in the excel
    sheet."""
    for index, row in teacher_df.iterrows():
        teacher_handle = row['teacher_handle']
        transcript_list = row['marked_transcripts'].split('\n')
        if row['wb_boolean'] is True:
            transcript_title = row['lesson_name'] + " -Whiteboard Used"
        else:
            transcript_title = row['lesson_name']
        ws.cell(row=1, column=index+1, value=transcript_title).font=Font(bold=True)
        ws.column_dimensions[get_column_letter(index+1)].width = int(70)
        for_line_in_transcript(transcript_list, teacher_handle, ws, index+1)
    

def for_line_in_transcript(transcript_list, teachername, ws, num):
    """Takes each lines from transcript lists and pastes it into a row 
    in the column. There is 1 column for each transcript, and then 1 row for
    each line. It also colors lines that have vocab found, or approriate phrase
    founds"""
    for row, line in enumerate(transcript_list, 0):
        active_cell = ws.cell(row=row+2, column=num)
        if '--VOCAB FOUND' in line:
            line = line.replace('--VOCAB FOUND',"")
            active_cell.font = Font(color="69a1e5",bold=True)
        if ' ~APPROPRIATE PHRASE' in line:
            line = line.replace('~APPROPRIATE PHRASE', "")
            active_cell.font = Font(color="cc79d1",bold=True)
        active_cell.alignment =  active_cell.alignment.copy(wrapText=True)
        active_cell.value = line


def create_stats_table(df, teacher_df, ws):
    stats_table = {
        'Labels':['Median','Max', "Min","Average"],
        'Teacher_FRT':[teacher_df.frt.median(), teacher_df.frt.max(), teacher_df.frt.min(), teacher_df.frt.mean()],
        'Teacher_ART':[teacher_df.art.median(), teacher_df.art.max(), teacher_df.art.min(), teacher_df.art.mean()],
        'Team_FRT':[df.frt.median(), df.frt.max(), df.frt.min(), df.frt.mean()],
        'Team_ART':[df.art.median(), df.art.max(), df.art.min(), df.art.mean()]}
    stats_table = pd.DataFrame(stats_table)
    rows = dataframe_to_rows(stats_table, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    for column in range(1,8):
        ws.column_dimensions[get_column_letter(column)].width = int(20)


def paste_time_stamps(teacher_df, ws):
    col = 8
    for transcript, item in enumerate(teacher_df.itertuples(), 1):
        time_stamp_title = "Transcript "+str(get_column_letter(transcript))
        ws.cell(row=1, column=col, value=time_stamp_title) 
        marked_timestamps = item.marked_timestamps
        response_df = rt_df(marked_timestamps, item.student_handle)
        for index, row in enumerate(response_df.itertuples(), 1):
            active_cell = ws.cell(row=index+2, column=col)
            active_cell.value = row.TimeStamps
            rt_cell = ws.cell(row=index+2, column=col+1)
            rt_cell.value = row.ResponseTimes
            if row.SFR:    
                active_cell.font = Font(color="228B22", bold=True)
                rt_cell.font = Font(color="228B22", bold=True)
            if row.TR is True:
                active_cell.font = Font(bold=True)
                rt_cell.font = Font(bold=True)
            if row.TFR is True:
                active_cell.font = Font(color=colors.BLUE, bold=True)
                rt_cell.font = Font(color=colors.BLUE, bold=True) 
        ws.column_dimensions[get_column_letter(col)].width = int(35)
        ws.column_dimensions[get_column_letter(col+1)].width = int(10)
        col = col+2
