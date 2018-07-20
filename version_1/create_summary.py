import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def create_summary(df):
    summary_book = Workbook()
    ws = summary_book.create_sheet("Team Summary", 0) 
    teacher_real_names = df.name
    unique_teacher_names = list(set(teacher_real_names))
    #create_titles(ws)
    summary = {'Name':[],
        'Wb':[],
        'vocab':[],
        'appropr':[],
        'SRL':[],
        'TRL':[],
        'SLS':[],
        'SLM':[],
    }   
    for teachername in unique_teacher_names:
        teacher_df = df[(df.name == teachername)]
        teacher_df = teacher_df.reset_index(drop=True)
        teacher_df.to_csv('Test.csv')
        summary['Name'].append(teachername)
        summary['Wb'].append(sum(teacher_df.wb_boolean))
        summary['vocab'].append(teacher_df.vocab_count.mean())
        summary['appropr'].append(teacher_df.approp_count.mean())
        summary['SRL'].append(None)
        summary['TRL'].append(None)
        summary['SLS'].append(teacher_df.session_length.mean())
        summary['SLM'].append((teacher_df.session_length.mean())/60)
    summary = pd.DataFrame(summary)   
    summary = summary.reset_index(drop=True)
    summary = summary[['Name', 'Wb', 'vocab', 'appropr', 'SRL', 'TRL', 'SLS', 'SLM']]
    rows = dataframe_to_rows(summary, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    summary_book.save('Test3.xlsx')
    
           
