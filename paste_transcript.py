from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import datetime

"""Patest the marked_lines column into each column in an excel sheet"""


def paste_response(row, rt_ws):
    """Finds the correct column to paste response time data into. Then 
    it drills down to the trans_df from the transcript df, which was zipped up
    earlier."""
    student_handle = row.student_handle
    trans_df = row.transcript_info
    trans_df = pd.DataFrame(trans_df)
    if rt_ws.cell(row=1, column=7).value is None:
        column = 7
    else:
        column = rt_ws.max_column+1
    title = 'Transcript '+(get_column_letter((column-4)/2))
    rt_ws.cell(row=1, column=column).value = title
    rt_ws.cell(row=1, column=column).font = Font(bold=True)    
    rt_ws.column_dimensions[get_column_letter(column)].width = int(35)
    rt_ws.column_dimensions[get_column_letter(column+1)].width = int(20)
    for_line_in_rt_ws(rt_ws, trans_df, column, student_handle)


def for_line_in_rt_ws(rt_ws, trans_df, column, student_handle):
    """Pastes data into the correct column. It then makes the first time a 
    student talks green, and a teachers first response to that green time, 
    in blue. After that all teacher responses are bolded"""
    FRT = 'Not Found'
    Student_First_Response = "Not Found"

    for row, line in enumerate(trans_df.rt_paste, 0):
        active_cell = rt_ws.cell(row=row+2, column=column+1)
        timestamp_cell = rt_ws.cell(row=row+2, column=column)
        time = line
        if "-TR" in str(time):
            time = int(time[:time.index('.0-TR')])
            time = str(datetime.timedelta(seconds=time))
            if FRT == 'Not Found':
                active_cell.font = Font(color=colors.BLUE,bold=True)
                timestamp_cell.font = Font(color=colors.BLUE,bold=True)
                FRT = 'Found'
            else:
                active_cell.font = Font(bold=True)
                timestamp_cell.font = Font(bold=True)
            active_cell.value = time
        else:
            if Student_First_Response == "Not Found" and student_handle.strip()==trans_df.Handle.loc[row]:
                active_cell.font = Font(color='00A563',bold=True)
                timestamp_cell.font = Font(color='00A563',bold=True)
                Student_First_Response = "Found"
            time = trans_df.rt.loc[row]
            #By making it a string, excel will automatically align it to the left of cell
            active_cell.value = (str(time)[6:]).strip()
        stamp = "  "+trans_df.Handle.loc[row]+" "+str(trans_df.Time_Stamps.loc[row])
        timestamp_cell.value = stamp
        
        
        

def paste_transcript(row, ws):
    """Takes the transcript row, and drills back down to trans_df and makes 
    it it's own df again. Then it finds the correct column to paste the 
    transcript in. It then creates the title of the transcript which is 
    the lesson name, and if the whiteboard boolean is true it pastes a 
    --Whiteboard Used at the end of the lesson name. This is placed in the
    first row in the column. It then passes the trans_df into for_line_in
    _transcript function"""
    wb_boolean = row.wb_boolean
    lesson_name = row.lesson_name
    trans_df = row.transcript_info
    item_number = str(row.item_number)
    trans_df = pd.DataFrame(trans_df)
    reason = row['reason']
    if ws.cell(row=1, column=1).value is None:
        column = 1
    else:
        column = ws.max_column+1
    if wb_boolean is True:
        transcript_title = str(lesson_name) +" - "+str(item_number)+" - "+str(reason)+ " -Whiteboard Used"
    else:
        transcript_title = str(lesson_name) +" - "+str(item_number)+" - "+str(reason)
    ws.cell(row=1, column=column, value=transcript_title).font=Font(bold=True)
    ws.column_dimensions[get_column_letter(column)].width = int(70)
    for_line_in_transcript(trans_df, ws, column)


def for_line_in_transcript(trans_df, ws, column):
    """Pastes each line of the transcript into it's own row. Then depending on
    if the indicators or there, does some kind of markup. For example if
     --VOCAB FOUND is in the line, then it makes that paticular cell's 
     font blue and bold."""
    for row, line in enumerate(trans_df.marked_lines, 0):
        active_cell = ws.cell(row=row+2, column=column)
        if '-- VOCAB FOUND' in line:
            line = line.replace('-- VOCAB FOUND',"")
            active_cell.font = Font(color="69a1e5",bold=True)
        if '--APPROP FOUND' in line:
            line = line.replace('--APPROP FOUND', "")
            active_cell.font = Font(color="cc79d1",bold=True)
        if '--MARK GREEN' in line:
            line = line.replace('--MARK GREEN', "")
            active_cell.font =  Font(color='00A563',bold=True)
        if "--GREY OUT" in line:
            line = line.replace('--GREY OUT', "")
            grayFill = PatternFill("solid", fgColor="efefef")
            active_cell.fill = grayFill
        active_cell.alignment =  active_cell.alignment.copy(wrapText=True) 
        active_cell.value = line
         
        