import pandas as pd
import os.path
import datetime
from openpyxl import Workbook
import numpy as np
from paste_transcript import paste_transcript
from paste_transcript import paste_response
from paste_kpi import paste_kpi
from datetime import datetime
import time


def teacher_df(unique_teacher_names, ytd, df, team_rt, team_frt, summary, lead_name, debug):
    """Creates a df for each teacher and sends it over to all of the paste
    functions to paste into excel"""
    for teachername in unique_teacher_names:
        print(teachername)
        row_in_ytd = ytd.loc[ytd['teacherName'] == teachername]
        teacherbook, ws, rt_ws = create_teacherbook()

        #Creates teacher_df, which is a df of just a single teacher transcripts.
        teacher_df = df[(df.name == teachername)]
        #teacher_df = teacher_df.sort_values('lesson_name')

        #Creates a list of teacher response times.
        teacher_rt = np.asarray([item for sublist in teacher_df.art.values for item in sublist])
        teacher_frt = np.asarray(teacher_df.frt.values.astype('timedelta64[s]'))
        #Pastes everything for each personal teacherbook.
        paste_kpi(teacher_rt, teacher_frt, team_rt, team_frt, rt_ws, row_in_ytd, teacherbook)
        teacher_df['visual_used'] = teacher_df.wb_boolean + teacher_df.has_drag_drop
        teacher_df['visual_used'] = teacher_df.visual_used.apply(lambda x: True if x >= 1 else False)

        teacher_df.apply(paste_transcript, args=(ws, ), axis=1)
        teacher_df.apply(paste_response, args=(rt_ws, ), axis=1)      
        #Saves the workbook as an excel doc, under teacher_sheets, using teachername in the title.
        summary = update_summary_data(teacher_df, teacher_rt, teacher_frt, summary, teachername, lead_name)
        
        mydate = datetime.now()
        month = mydate.strftime("%b")
        path = 'C:\\Users\kelly.richardson\OneDrive - Imagine Learning Inc\Reports\Transcript Reports'
        file_name = teachername+"_"+month+'-Transcript Report.xlsx'
        save_location = os.path.join(path,lead_name,teachername)
        if not os.path.isdir(save_location):
            os.makedirs (save_location)
        if debug is False:
           teacherbook.save(save_location+"/"+file_name) 
        else:
            teacherbook.save('teacher_sheets'+"/"+file_name)
    
    return summary


def create_teacherbook():
    """Sets up teacherbooks, creates worksheets, deletes default sheet"""
    teacherbook = Workbook()
    ws = teacherbook.create_sheet('Transcripts')
    rt_ws = teacherbook.create_sheet('ART-FRT')
    del teacherbook['Sheet']
    return teacherbook, ws, rt_ws


def update_summary_data(teacher_df, teacher_rt, teacher_frt, summary, teachername, lead_name):
    """Updates the Summary df which will be saved as a csv file and given to management"""
    frt = np.median(teacher_frt)
    art = np.median(teacher_rt)
    frt =  frt.item().total_seconds()
    art = art.item().total_seconds()
    teacher_summary = {
    'Name': [teachername],
    'Team': [lead_name],
    'FRT Median': [frt],
    'ART Median': [art],
    'Math Terms Per Session': [teacher_df.vocab_count.sum()/teacher_df.shape[0]],
    'Appropriate Phrase Per Session' : [teacher_df.approp_count.mean()/teacher_df.shape[0]],
    'Student Response Length': [np.median(teacher_df.avg_student_response_length)],
    'Teacher Response Length': [np.median(teacher_df.avg_teacher_response_length)],
    'Student to Teacher Exchange Ratio': [teacher_df.exchange_ratio.mean()],
    'Average Session Length(minutes)': [round((teacher_df.session_length_secs.mean())/float(60),2)],
    'Average Session Length(seconds)': [teacher_df.session_length_secs.mean()],
    'Visuals Count' : [teacher_df.visual_used.sum()],
    'Date': [datetime.today().strftime('%Y-%m-%d')]}
    teacher_summary = pd.DataFrame(teacher_summary)
    summary =  summary.append(teacher_summary, sort=False)
    return summary