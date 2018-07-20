from openpyxl import Workbook
import pandas as pd
import re
from marked_lines import create_marked_lines
from paste_transcript import paste_transcript
#from new_response_times import new_response_times

#New Approach: Deal with Transcript Individually and paste into workbook versus 
#trying to deal with it all at once.
def response_df(df):
    teacher_real_names = df.name
    unique_teacher_names = list(set(teacher_real_names))
    for teachername in unique_teacher_names:
        teacherbook = Workbook()
        trans_ws = teacherbook.create_sheet("Transcripts", 0)
        rt_ws = teacherbook.create_sheet("Response Time", 0) 
        #Creates a dataframe for each teacher.
        teacher_df = df[(df.name == teachername)]
        teacher_df = teacher_df.reset_index(drop=True)
        for index, row in teacher_df.iterrows():
            #Looks at each row for each teacher df. Works with one transcript at a time. 
            teach_handle = row['teacher_handle']
            stud_handle = row['student_handle']
            transcript = row['transcript'].split('\n')
            trans_df = create_trandsdf(transcript, teach_handle, stud_handle)
            trans_df = create_marked_lines(trans_df)
            paste_transcript(trans_ws, trans_df, row['wb_message_count'], row['lesson_name'])
        teacherbook.save('teacher_sheets/'+teachername+'.xlsx')


def create_trandsdf(transcript, teach_handle, stud_handle):
    """Creates the Transcript Dataframe, and finds the timestamp."""
    trans_df = {
        'Transcript':[], 
        'Time_Stamps':[],
        'Teacher_Bool':[],
        'Student_Bool':[],
        'Line_Char_Length':[]}
    for line in transcript:
        trans_df = fill_in_transdf(trans_df, line, teach_handle, stud_handle)
    trans_df = pd.DataFrame(trans_df)   
    trans_df = trans_df.reset_index(drop=True)
    return trans_df


def fill_in_transdf(trans_df, line, teach_handle, stud_handle):
    """Uses Logic to fill in the transcript dataframe"""
    time_stamp = create_timestamp(line)
    trans_df['Transcript'].append(line)
    trans_df['Time_Stamps'].append(time_stamp)
    trans_df['Line_Char_Length'].append(len(line))
    if teach_handle in time_stamp:
        trans_df['Teacher_Bool'].append(True)
        trans_df['Student_Bool'].append(False)
    elif stud_handle in time_stamp:
        trans_df['Teacher_Bool'].append(False)
        trans_df['Student_Bool'].append(True)
    else:
        trans_df['Teacher_Bool'].append(False)
        trans_df['Student_Bool'].append(False)
    return trans_df

def create_timestamp(line):
    """Creates a cleaned up timestamp"""
    time = line[:line.index('Z]:')+1]
    match = re.match(r"(.*)\[(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})", time)
    time_stamp = match.group(1)+" "+match.group(2)+" "+match.group(3)
    return time_stamp