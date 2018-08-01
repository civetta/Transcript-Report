import csv
import os
import datetime
from openpyxl import Workbook
from transcript_filter import get_filtered_teacher_transcripts, find_teacher_names
from cell_manipulator import remove_broken_transcripts,transcript_organizer, art_frt_paster, set_up_sheet, team_data_paste, copy_sheets, create_summary_page
from create_time_list import transcript_clean_up, art_frt_organizer, time_difference
from frt_art_calculator import session_length,find_rt, find_teacher_frt, find_teacher_ART, find_team_FRT, find_team_ART
from worksheet_setup import create_information_sheet
from copy_art import copy_past_kpi
from kpi_worksheet import gray_out,rt_paster,exchange_ratio,teacher_response_length,student_response_length,appropriate_phrase_count, vocab_phrase_count, create_boxplot
from dash_data import get_data
teacher_frt_art = {}
"""The following takes in 3 inputs. The file name, the number of transcripts, and the leads name which is used for naming file names"""
"""file_name = raw_input('File Name: ')
file_name = str(file_name)
transcript_number = raw_input('Number of Transcripts Per Teacher Youd Like to Return?:')
transcript_number = int(transcript_number)
team_lead_name = raw_input('Team Lead Name: ')
team_lead_name = str(team_lead_name)"""
file_name="CristenP"
transcript_number=20
team_lead_name="Angela"
#"Jeremy Shock","Rachel Adams","Jairo  Rios","Jill Szafranski","Kristin Donnelly","Caren Glowa",''
import datetime
mydate = datetime.datetime.now()
month=mydate.strftime("%B")
print month
"""The following code does 3 things. It pulls the csv file which must always be located under Program/ARTFRT.
Next is create a directory in that folder with Team_LeadName_CSV File Name
Finally it creates a final save for the Team_Lead which contains all of the information for each teacher. So Transcript + FRt/ART info + Team Info""" 
csv_file_name = os.path.join('C:\Users\kheyden\Documents\Program\ARTFRT', file_name + '.csv')
folder_name = team_lead_name
folder_location = os.path.join('C:\Users\kheyden\OneDrive - Imagine Learning\Reports\Transcript Reports', folder_name,"LEADBOOKS")
if not os.path.exists(folder_location):
    os.makedirs(folder_location)
final_save_name = os.path.join(folder_location, month+"_LEADBOOK_" + team_lead_name + '.xlsx')


"""The following takes the csv_file and returns a giant list of lists which we will use to iteracte"""
with open(csv_file_name, 'rb') as csv_file:
    reader = csv.reader(csv_file)
    csv_list = list(reader)

"""The following calls the Transcript filter, which taks our csv_list and returns a dictionary of teacher:transcript pairing.
And returns a list of all of the teacher names"""
teacher_transcript_dictionary = get_filtered_teacher_transcripts(csv_list, transcript_number)
teacher_names = find_teacher_names(csv_list)


def remove_sheet(wb):
    std = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)


def create_team_lead_workbook(teacher_names, teacher_transcript_dictionary):
    rt_dict={}
    approp_count_dict={}
    vocab_count_dict={}
    teacher_response_dict={}
    student_response_dict={}
    exchange_dict={}
    session_length_dict={}
    wb = Workbook()
    json=get_data()
    
    for i in teacher_names:
        a=0
        while a < len(json['ytdTeacher']):
            name = json['ytdTeacher'][a]['teacherName']
            if i==name:
                ytd=json['ytdTeacher'][a]['sessionCount']
                session_time=json['ytdTeacher'][a]['averageSessionTime']
            a=a+1
        print i
        session_time=(session_time)/float(60)
        teacher_transcript_dictionary[i]=remove_broken_transcripts(teacher_transcript_dictionary[i],transcript_number)
        teacher_transcript_dictionary[i]=sorted(teacher_transcript_dictionary[i])
        current_teacher_sheet = wb.create_sheet(i)
        transcript_organizer(teacher_transcript_dictionary[i], current_teacher_sheet, )

        """The following code blocks breaks the transcripts down to the time stamps and names. Identifies the FRT/ART. """
        time_list = transcript_clean_up(teacher_transcript_dictionary[i], i)
        rt_list = find_rt(time_list)
        teacher_average_frt = find_teacher_frt(rt_list)
        
        teacher_ART = find_teacher_ART(rt_list)
        teacher_frt_art[i] = rt_list
        print sum(teacher_average_frt)/len(teacher_average_frt)
        print sum(teacher_ART)/len(teacher_ART)
        print "\n"
        """The following code block pasts the teachers FRT/ART into a the worksheet"""
        #marked_time_list = identify_rt_frt(time_list)
        art_frt_teacher_sheet = wb.create_sheet(i+" FRT+ART")
        #copy_past_kpi(i,art_frt_teacher_sheet)
        set_up_sheet(art_frt_teacher_sheet)
        art_frt_paster(time_list, art_frt_teacher_sheet, teacher_ART, teacher_average_frt,ytd,session_time)
        average_session_length=session_length(time_list,transcript_number)
        session_length_dict[i]=average_session_length
        #kpi_worksheet=wb.create_sheet(i+" KPI")
        approp_count=appropriate_phrase_count(current_teacher_sheet,transcript_number)
        vocab_count=vocab_phrase_count(current_teacher_sheet,transcript_number)
        vocab_count_dict[i]=vocab_count
        approp_count_dict[i]=approp_count
        teach_response=teacher_response_length(current_teacher_sheet, transcript_number)
        teacher_response_dict[i]=teach_response
        student_response=student_response_length(current_teacher_sheet, transcript_number)
        student_response_dict[i]=student_response
        gray_out(current_teacher_sheet,transcript_number)
        ex_ratio=exchange_ratio(current_teacher_sheet,transcript_number)
        exchange_dict[i]=ex_ratio
        #mega_list=rt_paster(ex_ratio,teach_response,student_response,rt_list, kpi_worksheet,approp_count,vocab_count,transcript_number,i)
        
        flat_list=[]

        
        
        for sublist in rt_list:
            for item in sublist:
                h, m, s = item.split(':')
                item2 = int(h) * 3600 + int(m) * 60 + int(s)
                flat_list.append(item2)

        rt_dict[i]=flat_list
        
    
    team_frt = find_team_FRT(teacher_frt_art, teacher_names)
    team_art = find_team_ART(teacher_frt_art, teacher_names)


    
    for i in teacher_names:
        #create_boxplot(rt_dict,teacher_names,wb)
        ws = wb[i+" FRT+ART"]
        team_data_paste(ws, team_frt, team_art)
    summary = wb.create_sheet('Team Summary')
    create_summary_page(session_length_dict,teacher_names,wb,summary,rt_dict,vocab_count_dict,approp_count_dict,student_response_dict,teacher_response_dict,exchange_dict)
        
    remove_sheet(wb)
    wb.save(final_save_name)
    return wb
    

def create_individual_teacher_sheets(teacher_names,o_wb):
    for i in teacher_names:
        wb = Workbook()
        information_sheet = wb.create_sheet("Information")
        create_information_sheet(information_sheet)

        new_transcripts = wb.create_sheet(i)
        new_art_frt = wb.create_sheet(i +" FRT+ART")
        o_frt=o_wb[i +" FRT+ART"]
        o_trans=o_wb[i]
        copy_sheets(new_transcripts,new_art_frt,o_trans,o_frt)
        


        remove_sheet(wb)
        teacher_folder = os.path.join('C:\Users\kheyden\OneDrive - Imagine Learning\Reports\Transcript Reports', folder_name,i+" Transcript Reports")
        if not os.path.exists(teacher_folder):
            os.makedirs(teacher_folder)
        teacher_save_name = os.path.join(teacher_folder, i+"_"+month+"-Transcript Report"+ '.xlsx')
        wb.save(teacher_save_name)


wb=create_team_lead_workbook(teacher_names, teacher_transcript_dictionary)
#create_individual_teacher_sheets(teacher_names, wb) 
