from openpyxl.styles import Font
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter
def create_information_sheet(information):
    String0="Student First Response: The students first response after the Automated Message is colored in green. This is used to calculate the First Response Time"
    String1="FRT: FRT stands for First Response Time. This is the difference in time between the students first response after the automated message, and the teachers response to that very first message. The FRT is colored blue"
    String2="ART: ART stands for Average Response Time. This is the time in between a students response and the first time a teacher replies to that response. The ART is bold"
    information.cell(row=1,column=1,value=String0).font=Font(color="228B22",bold=True)
    information.cell(row=2,column=1,value=String1).font=Font(color="0000ff" ,bold=True)
    information.cell(row=3,column=1,value=String2).font=Font(bold=True)


    
