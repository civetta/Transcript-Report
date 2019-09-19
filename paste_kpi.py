from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from openpyxl.chart import BarChart, Series, Reference
"""Pastes the summary data in the KPI worksheet in the teacherbook. It includes
the team FRT, teacher FRT, teacher ART, and Team ART. Then it generates charts
for easy comparison"""


def paste_kpi(teacher_rt, teacher_frt, team_rt, team_frt, rt_ws, teacherbook,yeardata):
    """Creates a KPI dictionary with teacher stats,
     makes it into a dataframe, and then pastes the df into excel. Then it 
     calls ytd from the Admin Dashboard website"""
    print (yeardata[0])
    print (yeardata[1])
    print(("FRT: "+str(np.median(teacher_frt).astype(int))))
    print(("ART: "+str(np.median(teacher_rt).astype(int))))
    print (("YTD Session Taught: "+str(yeardata[1])))
    print ("")
    rt_ws.cell(row=2, column=1, value='FRT Median')
    rt_ws.cell(row=3, column=1, value='ART Median')
    rt_ws.cell(row=1, column=2, value='Teacher')
    rt_ws.cell(row=1, column=3, value='Team')
    rt_ws.cell(row=2, column=2, value=np.median(teacher_frt).astype(int))
    rt_ws.cell(row=2, column=3, value=np.median(team_frt).astype(int))
    rt_ws.cell(row=3, column=2, value=np.median(teacher_rt).astype(int))
    rt_ws.cell(row=3, column=3, value=np.median(team_rt).astype(int))
    rt_ws.column_dimensions['A'].width = int(35)
    create_box_chart(rt_ws)
    rt_ws.cell(row=6, column=1, value="YTD Session Taught")
    rt_ws.cell(row=6, column=2, value=int(yeardata[1]))
    rt_ws.cell(row=7, column=1, value="YTD Avg Session Length (minutes)")
    rt_ws.cell(row=7, column=2, value=round(float((yeardata[0])/60), 2))

def create_box_chart(rt_ws):
    """Uses openpyxl built in library to creates box charts of
     the data for easy reading"""
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Response Time Comparison"
    chart1.y_axis.title = 'Seconds'
    chart1.width = 12

    data = Reference(rt_ws, min_col=2, min_row=1, max_col=3, max_row=3)
    titles = Reference(rt_ws, min_col=1, min_row=2, max_row=3)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(titles)
    chart1.shape = 4
    rt_ws.add_chart(chart1, "A10")