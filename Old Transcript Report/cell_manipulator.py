# -*- coding: cp1252 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import alignment
import numpy as np
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


def remove_broken_transcripts(transcript_list,num):
    i = 0
    n = len(transcript_list)
    while i < n:
        transcript = transcript_list[i]
        new_list = transcript.splitlines()
        for line in range(len(new_list)):
            if "@ [" not in new_list[line]:
                print "FOUND------------------------------------------------------------------------------"
                transcript_list.remove(transcript)
                n = n - 1
                break
            else:
                i = i + 1
        
        
    print "passed function"          
    transcript_list=transcript_list[:num]
                
    return transcript_list
            

def transcript_organizer(transcript_list, ws):
    """This functions takes a list
    of transcipts and places each transcript in it's own column"""
    for i in transcript_list:
         ws.column_dimensions[get_column_letter(transcript_list.index(i)+1)].width  =  int(75)
         new_list = i.splitlines()
         x = 0
         while x<len(new_list):
             new_cell_trans=ws.cell(row = x+1,  column  =  transcript_list.index(i)+1,  value = new_list[x])
             if "LESSON SPLIT" in new_cell_trans.value:
                 cell_content=new_cell_trans.value
                 cell_content=cell_content.encode("utf8")
                 lesson_name=cell_content[(re.search("LESSON SPLIT",cell_content).start())+13:(re.search("LESSON SPLIT2",cell_content).start())]
                 ws.cell(row = x+1,  column  =  transcript_list.index(i)+1,  value = lesson_name).font=Font(bold=True)
                 
             new_cell_trans.alignment =  new_cell_trans.alignment.copy(wrapText=True)
             x = x+1
    return "hey"


def set_up_sheet(ws):
    ws.cell(row=1,column=5,value="Team ART")
    ws.cell(row=1,column=4,value="Team FRT")
    ws.cell(row=1, column=2,value="Teacher FRT")
    ws.cell(row=1, column=3,value="Teacher ART")
    ws.cell(row=1, column=1,value="Statistics")
    ws.cell(row=2, column=1,value="Median")
    ws.cell(row=3, column=1,value="Max")
    ws.cell(row=4, column=1,value="Min")
    ws.cell(row=5, column=1,value="Avg")
    
    s=1
    while s<125:
        ws.column_dimensions[get_column_letter(s)].width  =  int(30)
        s=s+1
    return ws


def art_frt_paster(time_list,ws,art,frt,ytd,session_length):
    ws.cell(row=2, column =2, value =np.median(frt,axis=0))
    ws.cell(row=3, column =2, value =frt[-1])
    ws.cell(row=4, column =2, value =frt[0])
    ws.cell(row=5, column =2, value =np.average(frt))
    ws.cell(row=2, column =3, value =np.median(art,axis=0))
    ws.cell(row=3, column =3, value =art[-1])
    ws.cell(row=4, column =3, value =art[0])
    ws.cell(row=5, column =3, value =np.average(art))
    ws.cell(row=7, column=1,value="YTD Students Taught")
    ws.cell(row=7, column=2, value=ytd)
    ws.cell(row=8, column=1,value="YTD Average Session Length")
    ws.cell(row=8, column=2, value=session_length)
    ws.column_dimensions['F'].width=30
    col=7
    for transcript in time_list:
        row=2
        ws.cell(row=1, column =col, value ="Transcript "+str((get_column_letter(time_list.index(transcript)+1))))
        ws.cell(row=1,column=col).font=Font(bold=True)
        for transaction in transcript:
            cell1=ws.cell(row=row, column=col)
            cell2=ws.cell(row=row,column=col+1)
            if 'GREEN' in transaction[0]:
                transaction[0]=transaction[0][transaction[0].index(" "):]
                cell1.font=Font(color="228B22",bold=True)
                cell2.font=Font(color="228B22",bold=True)
            if 'BLUE' in transaction[0]:
                transaction[0]=transaction[0][transaction[0].index(" "):]
                cell1.font=Font(color=colors.BLUE,bold=True)
                cell2.font=Font(color=colors.BLUE,bold=True)
            if 'BOLD' in transaction[0]:
                transaction[0]=transaction[0][transaction[0].index(" "):]
                cell1.font=Font(bold=True)
                cell2.font=Font(bold=True)
            ws.cell(row=row, column =col, value =transaction[0])
            ws.cell(row=row, column =col+1, value =transaction[1])
            row=row+1
        col=col+2
    create_bar_chart(ws)
    return ws




def copy_sheets(new_transcripts,new_times,o_trans,o_times):
    col=1
    while col<120:
        new_times.column_dimensions[get_column_letter(col)].width = int(30)
        new_transcripts.column_dimensions[get_column_letter(col)].width = int(75)
        r=1
        fr=1
        while r<70:
            o_cell_trans = o_trans.cell(row=r,column=col)
            new_cell_trans= new_transcripts.cell(row=r,column=col, value=o_cell_trans.value)
            new_cell_trans.alignment =  new_cell_trans.alignment.copy(wrapText=True)
            if o_cell_trans.font==Font(color="cc79d1",bold=True):
                new_cell_trans.font=Font(color="cc79d1",bold=True)
            if o_cell_trans.font==Font(color="69a1e5",bold=True):
                new_cell_trans.font=Font(color="69a1e5",bold=True)
            if o_cell_trans.font==Font(bold=True):
                new_cell_trans.font=Font(bold=True)
            if o_cell_trans.font==Font(color="3f8e1d",bold=True):
                new_cell_trans.font=Font(color="3f8e1d",bold=True)
            if o_cell_trans.fill==PatternFill("solid", fgColor="efefef"):
                new_cell_trans.fill=PatternFill("solid", fgColor="efefef")
            o_cell_times= o_times.cell(row=r,column=col)
            new_cell_times= new_times.cell(row=r,column=col,value=o_cell_times.value)
            if o_cell_times.font==Font(color="228B22",bold=True):
                new_cell_times.font=Font(color="228B22",bold=True)
            if o_cell_times.font==Font(color=colors.BLUE,bold=True):
                new_cell_times.font=Font(color=colors.BLUE,bold=True)
            if o_cell_times.font==Font(bold=True):
                new_cell_times.font=Font(bold=True)
            
            

            
            r=r+1
        col=col+1
    create_bar_chart(new_times)

def create_bar_chart(ws):
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "FRT/ART"
    chart1.y_axis.title = 'Time (in seconds)'
    chart1.x_axis.title = ''

    data = Reference(ws, min_col=2, min_row=1, max_row=2, max_col=5)
    cats = Reference(ws, min_col=2, min_row=1, max_row=1, max_col=5)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "A10")


def create_summary_page(session_length,teacher_names,wb,summary,rt_dict,vocab_count_dict,approp_count_dict,student_response_dict,teacher_response_dict,exchange_dict):
    ws=wb[teacher_names[0] +" FRT+ART"]
    summary.column_dimensions[get_column_letter(1)].width = int(30)
    summary.cell(row=1,column=1,value="Name")
    summary.column_dimensions[get_column_letter(2)].width = int(30)
    summary.cell(row=1,column=2,value="FRT Median")
    summary.column_dimensions[get_column_letter(3)].width = int(30)
    summary.cell(row=1,column=3,value="ART Median")
    summary.column_dimensions[get_column_letter(4)].width = int(30)
    summary.cell(row=1,column=4,value="Average Vocab Count per Session")
    summary.column_dimensions[get_column_letter(5)].width = int(30)
    summary.cell(row=1,column=5,value="Average Appropriate Phrase Count per Session")
    summary.column_dimensions[get_column_letter(6)].width = int(30)
    summary.cell(row=1,column=6,value="Average Student Response Length")
    summary.column_dimensions[get_column_letter(7)].width = int(30)
    summary.cell(row=1,column=7,value="Average Teacher Response Length")
    summary.column_dimensions[get_column_letter(8)].width = int(30)
    summary.cell(row=1,column=8,value="Student to Teacher Ratio")
    summary.column_dimensions[get_column_letter(9)].width = int(30)
    summary.cell(row=1,column=9,value="Average Session Length")
    
    for i in teacher_names:
        ws=wb[i +" FRT+ART"]
        r=teacher_names.index(i)+2
        summary.cell(row=r,column=1,value=i)
        summary.cell(row=r,column=2,value=ws.cell(row=2, column=2).value)
        summary.cell(row=r,column=3,value=ws.cell(row=2, column=3).value)
        summary.cell(row=r,column=4,value=vocab_count_dict[i])
        summary.cell(row=r,column=5,value=approp_count_dict[i])
        summary.cell(row=r,column=6,value=np.average(student_response_dict[i]))
        summary.cell(row=r,column=7,value=np.average(teacher_response_dict[i]))
        summary.cell(row=r,column=8,value=exchange_dict[i])
        summary.cell(row=r,column=9,value=session_length[i])


        
            
        

def team_data_paste(ws,frt,art):
    ws.cell(row=2, column =4, value =np.median(frt,axis=0))
    ws.cell(row=3, column =4, value =frt[-1])
    ws.cell(row=4, column =4, value =frt[0])
    ws.cell(row=5, column =4, value =np.average(frt))
    ws.cell(row=2, column =5, value =np.median(art,axis=0))
    ws.cell(row=5, column =5, value =np.average(art))
    ws.cell(row=3, column =5, value =art[-1])
    ws.cell(row=4, column =5, value =art[0])
