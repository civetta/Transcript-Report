import openpyxl
from openpyxl.utils import get_column_letter
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import colors
import numpy as np
from openpyxl.chart import BarChart, Series, Reference
import matplotlib as mpl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import LineChart,Reference

def copy_past_kpi(i,kpiws):
    wb = load_workbook('ARTFRT.xlsx')
    ws = wb.worksheets[0]
    name_column=1
    row_count=29
    kpiws.cell(row=28,column=1).value='Month'
    kpiws.cell(row=28,column=2).value='FRT Median'
    kpiws.cell(row=28,column=3).value='ART Median'
    for row in range(1,ws.max_row+1):
        if ws.cell(row=row,column=name_column).value == i:
            while row_count<29+9:
                kpiws.cell(row=row_count,column=1).value=ws.cell(row=1,column=name_column).value
                kpiws.cell(row=row_count,column=2).value=ws.cell(row=row,column=name_column+1).value
                kpiws.cell(row=row_count,column=3).value=ws.cell(row=row,column=name_column+2).value
                name_column=name_column+4
                row_count=row_count+1
    c1 = LineChart()
    c1.title = "FRT/ART Over Time"
    c1.style = 13
    c1.y_axis.title = 'Seconds'
    c1.x_axis.title = 'Months'
    data = Reference(ws, min_col=1, min_row=28, max_col=3, max_row=36)
    c1.add_data(data, titles_from_data=True)
    ws.add_chart(c1, "D29")
