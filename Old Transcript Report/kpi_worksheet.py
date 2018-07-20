# -*- coding: cp1252 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import colors
import numpy as np
from openpyxl.chart import BarChart, Series, Reference
import matplotlib as mpl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

## agg backend is used to create plot as a .png file
mpl.use('agg')

import matplotlib.pyplot as plt 


def rt_paster(exchange_ratio,teacher_response,student_response,rt_list, ws,approp_count,vocab_count,num,teacher_name):
    ws.cell(row=1, column =1, value ="KPI")
    ws.cell(row=2, column =1, value ="FRT")
    ws.cell(row=3, column =1, value ="ART")
    ws.cell(row=6, column =1, value ="Appropriate Phrase to Transcript Ratio:")
    ws.cell(row=6, column =2, value =approp_count/num)
    ws.cell(row=7, column =1, value ="Vocab Phrases to Transcript Ratio:")
    ws.cell(row=7, column =2, value =vocab_count/num)
    ws.cell(row=8, column =1, value ="Teacher to Student Exchange Ratio:")
    ws.cell(row=8, column =2, value =exchange_ratio)
    for i in rt_list:
        standard_form=[]
        for a in i:
            h, m, s = a.split(':')
            standard_form.append(int(h) * 3600 + int(m) * 60 + int(s))
        ws.cell(row=1, column=rt_list.index(i)+3,value="Transcript"+ str(rt_list.index(i)+1))
        ws.cell(row=2, column=rt_list.index(i)+3,value=standard_form[0])
        ws.cell(row=3, column=rt_list.index(i)+3,value=(sum(standard_form) / float(len(standard_form))))




def teacher_response_length(ws, num):
    i=1
    response=[]
    while i<= num:
        r=2
        row_max=ws.max_row
        while r<row_max:
            cell_content=ws.cell(row=r, column=i).value
            if cell_content:
                name = cell_content[0:cell_content.index('@')]
                if re.search('Mrs|Ms|Miss|Mr',name):
                   response.append(len(cell_content[cell_content.index('Z]:')+3:]))
            r=r+1
        i=i+1
    return response

def student_response_length(ws, num):
    i=1
    response=[]
    while i<= num:
        r=2
        row_max=ws.max_row
        while r<row_max:
            current_cell=ws.cell(row=r,column=i)
            cell_content=current_cell.value
            if cell_content:
                name = cell_content[0:cell_content.index('@')]
                if re.search('^Mrs|^Ms|^Miss|^Mr|Server Notice',name)<=-1:
                   response.append(len(cell_content[cell_content.index('Z]:')+3:]))
            r=r+1
        i=i+1
    return response
			


def exchange_ratio(ws, num):
    i=1
    teacher_count=0
    student_count=0
    while i<= num:
        r=2
        row_max=ws.max_row
        while r<row_max:
            cell_content=ws.cell(row=r, column=i).value
            if cell_content:
                name = cell_content[0:cell_content.index('@')]
                if re.search('Mrs|Ms|Miss|Mr',name):
                    teacher_count=teacher_count+1
                    
                    
                else:
                    student_count=student_count+1
            r=r+1
        i=i+1
    return teacher_count/float(student_count)
            
    
def create_boxplot(rt_dict,teacher_name,wb):
    mega=create_mega_list(rt_dict,teacher_name)
    
    for i in teacher_name:
        ws=wb[i+" KPI"]
        # Create the boxplot
        lister=[mega,rt_dict[i]]
        # Create a figure instance
        fig = plt.figure(1, figsize=(9, 6))

        # Create an axes instance
        ax = fig.add_subplot(111)

        # Create the boxplot
        bp = ax.boxplot(mega)

        # Save the figure
        fig.savefig('fig1.png', bbox_inches='tight')
    
        img = Image(i+".png")
        ws.add_image(img, 'A1')


def create_mega_list(rt_dict,teacher_name):
    mega_list=[]
    for i in teacher_name:
        for a in rt_dict[i]:
            mega_list.append(a)
    return sorted(mega_list)
def column_width(number_of_transcripts,ws):
    num=number_of_transcripts+6
    i=1
    while i<num+7:
        ws.column_dimensions[get_column_letter(i)].width  =  int(20)
        i=i+1


def appropriate_phrase_count(ws,num):
    approp=["See if you can use what we did to try this on your own.","You are doing a great job! Time to work on your own.","I think you have it from here! Keep up the awesome work!","I have not heard from you in a few minutes.","Please come back if you need more help. Have a nice day!","I see why you might think that, but let's try this another way.","Great try. Let's see if that answer works!","Not quite, but good effort! Let's work together to figure this out.","Your choice of words does not meet our chat rules. Please be respectful. Would you like my help with math?","I would like to help you. Are you ready to work on this together?","I am sorry that you do not want to work with me. Please come back when you are ready.","It is okay if you do not know! We can work together to find out.","Please give it a try! It is okay to make mistakes.","That is okay. Would you be willing to try?","It's ok to not know things sometimes, but please try so that I may help you here :) What do you think?","It really helps to read or listen to the problem carefully. Are you able to do that?","What is one thing you remember from the problem?","Did you have a chance to check out your problem?","Let's check it out together! What does it ask you to find?","I am sorry for the delay. Thank you for your patience!","Please give me a moment to look over your work. We will get started soon!","Sorry you had to wait! Let's get started.","I'm sorry for the wait. I was helping another student. Let's get started!","Welcome! Let's look over your problem again.","Hi! Would you like me to type to you, or talk to you?","Hi! Do you have a headset or speakers so that I can talk to you?","I am sorry, we must not have a good connection today. Let's type.","I don't think you are able to hear me, but we can type!","Your device does not work with my mic for some reason. Let's type!","I am not able to switch your teacher. But, I am happy to help you!","I am sorry. We cannot switch your teacher. I am happy to help you!","I am not able to switch your teacher, but I am happy to help! Would you be willing to work with me?","Please do not share personal information on the internet. Let's work on math!","I can only help you here. You can always talk to our teachers about your math questions here!","That is personal information. Let's just stick to math!","I can only help you on our site. You can always talk to our teachers about your math questions here!","We are going to practice with one of the parts in your problem, then after that you can try it without me :) ready?","That is a very serious thing to say and I am sad to hear that you are in pain. I know that this can be really hard to deal with.","I have found it helpful to talk to someone when I am feeling sad or in pain. Have you tried talking to your teacher or a counselor at school about this?","It's important to have people around you that know what is happening. I will let someone at your school know that you'd like to talk.","Please think about contacting someone at one of these anonymous helplines: 1-800-784-2433 **or** 1-800-273-8255","*Link to Suicide and Other Threats doc*","Are you okay?","I will alert your teacher. What is happening?","Have you talked to an adult about this?","I am sad to hear that you are dealing with this. I will let someone know for you.","Let's talk about your math problem first :)","The whiteboard is really cool, but the teacher will decide when to use it :)","Let's talk about your math problem first and then decide if the teacher needs to use it :)","Thank you for the suggestion! It is helpful to hear other strategies that you are using in your classroom. Would you like me to conitinue the session with","We are always adding new content & revising existing content, so your feedback is really helpful!","Thank you for pointing that out! I will share your concern with our curriculum department.","Thank you for sharing your concern! You can also contact support@thinkthroughmath.com. But, I can report this issue if you prefer.","There are multiple ways to solve any math problem. It is important to us that we dont prescribe methods, but rather helps students understand how to approach the math concept or idea.","Since this is an instructional product, one of our goals is to introduce and teach new concepts. Therefore, we provide lots of supports for our students, such as us teachers! I would be happy to help your student understand this concept.","Students work in the Guided Learning and Practice activity before they get to the quiz. In both activities, students work through the problem until they get the correct answer. Also, if students are not successful, they will receive some remediation lessons and then will be given the opportunity to take the lesson again.","It would be best if your concerns were reported to Customer Support. Would you like me to report for you? Otherwise, please feel free to email support@thinkthroughmath.com. We really appreciated this type of feedback.","Let's take a look at an example together.","The teacher here can see the question on which the student is working, as well as the answer choices and images. We use the feedback and visuals that the student has received, as well as simpler examples or other approaches to briefly explain concepts and help guide the student through the problem solving process. We can also use an interactive whiteboard to work on the bigger conceptual ideas. We can use pictures and diagrams to help the student so that when we're done, they can accomplish their problem on their own.","What does this problem ask you to find?","I cannot speak in Spanish, but I can try to type in Spanish. Would you like me to do that for you?","Your teacher has **not** indicated that you need help in Spanish. Would you like me to ask your teacher to change this for you?"]
    col=1
    vocab_count=0
    whiteboard_count=0
    while col<=num:
        row=2
        while row<ws.max_row:
            row=row+1
            current_cell= ws.cell(row =row,  column  =col)
            if current_cell.value:
                cell=current_cell.value
                for a in approp:                
                    if cell.find(a)>-1:
                        vocab_count=vocab_count+1
                        current_cell.font=Font(color="cc79d1",bold=True)
                        
        col=col+1
    return vocab_count/float(num)

def gray_out(ws,num):
    col=1
    while col<=num:
        row=2
        maxer=ws.max_row
        while row<maxer:
            row1=row
            row2=row+1
            row3=row+2
            current_cell1= ws.cell(row =row1,  column  =col)
            current_cell2= ws.cell(row =row2,  column  =col)
            current_cell3= ws.cell(row =row3,  column  =col)
            cell1=current_cell1.value
            cell2=current_cell2.value
            cell3=current_cell3.value
            if cell3:
                name1 = cell1[0:cell1.index('@')]
                name2 = cell2[0:cell2.index('@')]
                name3 = cell3[0:cell3.index('@')]
                if re.search('Mrs|Ms|Miss|Mr',name1) and re.search('Mrs|Ms|Miss|Mr',name2)and re.search('Mrs|Ms|Miss|Mr',name3) :
                        grayFill=PatternFill("solid", fgColor="efefef")
                        current_cell1.fill = grayFill
                        current_cell2.fill = grayFill
                        current_cell3.fill = grayFill
            row=row+1
        col=col+1
        
                    
            

def vocab_phrase_count(ws,num):
    vocab=['Absolute Value', 'Acute Angle', 'Addend', 'Addition', 'Additive Inverses', 'adjacent', 'Algorithm', 'Alternate Exterior Angles', 'Alternate Interior Angles', 'Angle', 'Approximately', 'Arc', 'Area', 'Area Model Of Multiplication', 'Arithmetic Sequence', 'Array', 'Associative Property', 'Associative Property Of Addition', 'Associative Property Of Multiplication', 'Asymptote', 'Average', 'Average Rate Of Change', 'Bar Chart', 'Bar Graph ', 'Base ', 'Base Ten', 'Base-10 ', 'Benchmark', 'Benchmark Fractions', 'Biconditional Statement', 'Binomial', 'Bisect', 'Box-And-Whisker Plot', 'Calculate', 'Central Angle', 'Chord', 'Circumference', 'Circumscribed Angle', 'Circumscribed Figure', 'Classify', 'Coefficient', 'Column', 'Combine', 'common denominator', 'Common Factor', 'Common Multiple', 'Common Ratio', 'Commutative Property', 'Complementary Angles', 'Completing The Square', 'composite figure', 'Composite Number', 'compound interest', 'compound probability', 'Concave', 'Conditional Statement', 'Cone', 'congruent', 'Congruent Figures', 'Conjecture', 'Constant Term', 'Contrapositive', 'Converse ', 'Coordinate Pair', 'Coordinate Plane', 'Correlation', 'Corresponding Angles', 'Corresponding Sides', 'Cosine Ratio', 'Cross Section', 'Cube ', 'Customary Measurement System', 'Cylinder', 'Data', 'Decimal', 'Decimal Approximation', 'Decimal Grid', 'Decrease', 'Degree', 'Delta', 'Denominator', 'Density', 'Diagonal', 'Diagram', 'Diameter', 'Difference', 'Digit', 'Dilation', 'Dimensions', 'direct variation', 'Directly Proportional', 'Distance', 'Distributive Property', 'Dividend', 'Division', 'Divisor', 'Domain', 'Dot Plot', 'Double', 'Edge', 'Elapsed Time', 'Endpoint', 'Equally Likely Events', 'Equation', 'Equiangular', 'Equidistant', 'Equilateral Triangle', 'equivalent', 'Equivalent Expressions', 'Equivalent Fractions', 'Equivalent Ratios', 'Estimate', 'estimation', 'Evaluate', 'Excluded Values', 'Expanded Exponential Form', 'Expanded Form ', 'Experimental Probability', 'Exponent', 'Exponent ', 'Exponential Decay', 'Exponential Growth', 'Expression', 'Exterior Angle', 'Fact Family', 'Factor', 'factor pair', 'Favorable Outcome', 'Figure', 'First Quartile', 'Five-Number Summary', 'Formula', 'Fraction', 'Frequency', 'Function', 'Gallon', 'Geometric Sequence', 'Graph', 'Greater Than', 'Greatest Common Factor', 'Group', 'Growth Factor', 'Half-Plane', 'Halves', 'Hand-Span', 'Height', 'Histogram', 'Horizontal', 'Hypotenuse', 'Identity Property Of Addition', 'Identity Property Of Multiplication', 'Imaginary Number', 'Improper Fraction', 'Increase', 'Inequality', 'Inscribed Angle', 'Inscribed Figure', 'Integers', 'Integers', 'Interest', 'Interior Angle', 'Interquartile Range', 'Intersection', 'Interval', 'inverse', 'Inverse Of A Conditional Statement', 'Inverse Operations', 'Inverse Relationship', 'Inversely Proportional', 'Irrational Number', 'Isosceles Triangle', 'Lateral Surface Area', 'Law Of Large Numbers', 'Least Common Denominator', 'Least Common Multiple', 'Legs', 'Length', 'Less Than', 'Like Terms', 'Line', 'Line Of Best Fit', 'Line Of Symmetry', 'Line Plot', 'Line Segment', 'Linear Pair', 'Linear Relationship', 'Location', 'Logarithm', 'Long Division', 'Lower Quartile', 'Mass', 'Mathematical Model', 'mean absolute deviation', 'Measure', 'Median', 'Meter', 'Metric System', 'Mixed Number', 'Mode', 'Model', 'Multiple', 'Multiplication', 'Multiplicative Identity', 'Multiplicative Inverses', 'Negative Number', 'Net', 'Normal Distribution', 'Number Line', 'Number System', 'Numerator', 'Obtuse Angle', 'odd', 'Operation', 'Opposites', 'Order Of Operations', 'Origin', 'Outcome', 'Parabola', 'Parallel Lines', 'Parallelogram', 'Parentheses', 'Pattern', 'Percent', 'Perfect Square', 'Perimeter', 'Perpendicular', 'Pi ', 'Place Value', 'Plane', 'point-slope form', 'Polygon', 'Polygon', 'Polynomial', 'Positive Number', 'Powers Of Ten', 'Prime Factorization', 'Prime Number', 'Principal', 'Prism', 'Probability', 'Product', 'profit', 'Proof', 'Proportion', 'proportional relationship', 'Pyramid', 'Pythagorean Theorem', 'Quadrants', 'Quadratic', 'Quadratic Formula', 'Quadrilateral', 'quartile', 'Quotient', 'Radian', 'Radical Expression', 'Radius', 'Range', 'Range Of A Function', 'Rate', 'Ratio', 'Rational Number', 'Ray', 'Reasonable', 'Reciprocal', 'Rectangle', 'Rectangular Prism', 'recursive formula', 'Reflection', 'Reflection Symmetry', 'Reflexive Property Of Equality', 'regroup', 'Regrouped', 'Regrouping', 'Regular Polygon', 'related facts', 'Remainder', 'Repeating Decimal', 'Represent', 'Rhombus', 'Right Angle', 'Right Prism', 'Right Triangle', 'Rigid Transformation', 'Rise', 'Roots Of A Function', 'Rotation', 'Rotational Symmetry', 'Round', 'Row', 'Same Side Exterior Angles', 'Same Side Interior Angles', 'Sample Space', 'Scale ', 'Scale Factor', 'Scalene Triangle', 'Scatter Plot', 'Section', 'Sector', 'Sequence', 'Shaded', 'Shape', 'Side Of A Polygon', 'Similar', 'Similar Figures', 'simple interest', 'Simplest Form ', 'Sine Ratio', 'Situation', 'Slope', 'Slope-Intercept Form Of An Equation', 'Sphere', 'Spread', 'Square', 'Square Number', 'Square Root', 'Square Unit', 'Standard Deviation', 'standard form', 'Stem-And-Leaf Plot', 'Straight Angle', 'Substitution', 'Subtend', 'Subtraction', 'Sum', 'Supplementary Angles', 'Surface Area', 'Symmetry', 'System Of Linear Equations', 'Table', 'Tally Marks', 'Tangent Line', 'Tangent Ratio', 'Term', 'Terminating Decimal', 'Theoretical Probability', 'Third Quartile', 'Transformation', 'Transformation Rule', 'Transitive Property Of Equality', 'Translation', 'Transversal', 'Trapezoid', 'Tree Diagram', 'Trial', 'Trigonometric Ratio', 'Trinomial', 'Unequal', 'Unit', 'Unit Cube', 'Unit Fraction', 'Unit Rate', 'Upper Quartile', 'value', 'Variable', 'Vertex', 'Vertical', 'Vertical Angle', 'Volume', 'Weight', 'Whole', 'Whole Numbers', 'word form', 'X-Axis', 'X-Intercept', 'Y-Axis', 'Y-Intercept', 'Zero Property', 'Zeros Of A Function', 'Add', 'Subtract', 'Multiply ', 'Divide', 'Convex', 'Overestimate', 'Equal', 'Equally', 'Total', 'Face']
    vocab=sorted(vocab,key=len)
    vocab=vocab[::-1]
    unique_vocab=[]
    col=1
    vocab_count=0
    whiteboard_count=0
    while col<=num:
        row=2
        while row<ws.max_row:
            current_cell= ws.cell(row =row,  column  =col)
            if current_cell.value:
                cell=current_cell.value
                try:
                    name = cell[0:cell.index('@')]
                except:
                    print cell
                if re.search('Mrs|Ms|Miss|Mr',name):
                    vocab_str=""
                    transcript=cell[cell.index('Z]:')+3:]        
                    for a in vocab:
                        rvocab='(^|\s|\W)'+a+'($|\s|\W|s)'
                        if re.search(rvocab,transcript,re.IGNORECASE)>-1:      
                            transcript=transcript[:re.search(rvocab,transcript,re.IGNORECASE).start()]+transcript[re.search(rvocab,transcript,re.IGNORECASE).start()+len(a):]
                            vocab_count=vocab_count+1
                            unique_vocab.append(a)
                            vocab_str=vocab_str+"-"+a
                            current_cell.value=cell +'   --"'+vocab_str+'"--'
                            current_cell.font=Font(color="69a1e5",bold=True)
                    if re.search("whiteboard",transcript,re.IGNORECASE)>-1:
                        current_cell.font=Font(color="3f8e1d",bold=True)
                    if re.search("draw",transcript,re.IGNORECASE)>-1:
                        current_cell.font=Font(color="3f8e1d",bold=True)
                    if "[Image]" in transcript:
                        current_cell.font=Font(color="3f8e1d",bold=True)
                        whiteboard_count=whiteboard_count+1
            row=row+1
        col=col+1
    
    return vocab_count/float(num)
